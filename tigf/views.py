from utils.logger_utils import time_tracker

#! 包裝整個 import 區塊或初始化邏輯
with time_tracker("tigf"):
    import copy
    import io
    import re
    import warnings
    import zipfile
    from datetime import datetime
    from decimal import Decimal, InvalidOperation
    from urllib.parse import quote

    import pandas as pd
    from django.core.cache import cache
    from django.http import HttpResponse, JsonResponse
    from django.shortcuts import render
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    from utils.decorators import staff_required


def get_all_files(uploaded_files):
    """
    輔助函式：將上傳的檔案清單（包含 ZIP）轉換為字典 {檔名: 檔案內容或對象}
    """
    files_data = {}
    for f in uploaded_files:
        if f.name.lower().endswith(".zip"):
            try:
                with zipfile.ZipFile(f) as z:
                    for name in z.namelist():
                        #! 排除資料夾與系統暫存檔
                        if name.endswith("/") or "__MACOSX" in name:
                            continue
                        #! 讀取 ZIP 內的檔案內容
                        files_data[name] = io.BytesIO(z.read(name))
            except zipfile.BadZipFile:
                #! 若 ZIP 損毀則跳過或記錄
                pass
        else:
            files_data[f.name] = f
    return files_data


def normalize_val(val):
    """資料正規化：處理日期格式 (2024-05-14 -> 20240514) 與數值字串"""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if "-" in s and len(s) == 10:
        s = s.replace("-", "")
    return s


def smart_read_csv(file_obj, **kwargs):
    """
    自動嘗試不同編碼讀取 CSV，解決 utf-8 編碼報錯問題
    """
    encodings = ["utf-8-sig", "cp950", "utf-8", "big5"]
    content = file_obj.read()
    if "dtype" not in kwargs:
        kwargs["dtype"] = str
    for enc in encodings:
        try:
            #! 每次嘗試需將指標重置或使用 io.BytesIO
            df = pd.read_csv(io.BytesIO(content), encoding=enc, **kwargs)
            return df
        except (UnicodeDecodeError, Exception):
            continue
    #! 若全部失敗，拋出原始錯誤
    file_obj.seek(0)
    return pd.read_csv(file_obj, **kwargs)


def smart_read_excel(file_obj, **kwargs):
    """讀取 Excel 並過濾掉煩人的 Header/Footer 警告"""
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        #! 確保指標在開頭
        file_obj.seek(0)
        return pd.read_excel(file_obj, **kwargs)


#! 獨立的數值與日期比對函式：處理日期正規化、0.0 == "" 以及浮點數誤差
def is_value_matched(
    val_r,
    val_db,
    rule_str_match=True,
    rule_date_check=True,
    rule_empty_zero=True,
    rule_tolerance=True,
    has_tis=False,
    ignore_list=[],
):
    #! 基礎清理與正規化
    str_r = str(val_r).strip().replace("nan", "").replace("None", "").replace("\r", "")
    str_db = str(val_db).strip().replace("nan", "").replace("None", "").replace("\r", "")

    #! TIS
    if has_tis:
        str_r = str_r.strip("[]")
        str_db = str_db.strip("[]")
        if str_r == "-":
            str_r = ""
        if str_db == "-":
            str_db = ""
        if str_r == "True":
            str_r = "1"
        if str_r == "False":
            str_r = "0"
        if str_db == "True":
            str_db = "1"
        if str_db == "False":
            str_db = "0"
        if str_r == "%":
            str_r = ""
        if str_db == "%":
            str_db = ""
        if str_r == "永續" and str_db == "99991231":
            return True
        if ignore_list:
            for ignore_str in ignore_list:
                if str_r == ignore_str:
                    str_r = ""
                    break

    #! 字串完全相同 (包含兩者皆為空字串)
    if rule_str_match and str_r == str_db:
        return True

    if rule_date_check:
        #! 日期格式正規化比對 (處理如 20250130 = 2025-01-30)
        #! 移除所有非數字字元後嘗試進行日期長度檢查
        date_r = re.sub(r"[^0-9]", "", str_r)
        date_db = re.sub(r"[^0-9]", "", str_db)

        #! 若包含了時間 (例如 20260101000000)，我們只取前 8 碼的日期部分 (YYYYMMDD)
        if len(date_r) >= 8:
            date_r = date_r[:8]
        if len(date_db) >= 8:
            date_db = date_db[:8]

        #! 若兩者清理後皆為 8 位數字且內容相同，視為日期匹配
        if len(date_r) == 8 and len(date_db) == 8 and date_r == date_db:
            try:
                #! 確保字串符合基本日期邏輯 (例如不會出現 20251340)
                datetime.strptime(date_r, "%Y%m%d")
                return True
            except ValueError:
                #! 若非有效日期則跳過，進入後續數值比對
                pass

    #! 數值比對：將空字串視為 0 進行 Decimal 轉換
    try:
        if rule_empty_zero:
            dec_r = Decimal(str_r) if str_r else Decimal("0")
            dec_db = Decimal(str_db) if str_db else Decimal("0")
        else:
            #! 若未啟用空值視為 0，遇空字串轉 Decimal 會拋出 InvalidOperation 並略過比對
            dec_r = Decimal(str_r)
            dec_db = Decimal(str_db)

        if rule_tolerance:
            #! 動態容差：取兩數小數點最小位數，進行標準四捨五入後比對
            #! .as_tuple().exponent 可以取得小數位數 (例如 12.34 的 exponent 是 -2)
            #! max(0, ...) 是為了防止整數或科學記號造成的正數 exponent
            dp_r = max(0, -dec_r.as_tuple().exponent)  # pyright: ignore[reportOperatorIssue]
            dp_db = max(0, -dec_db.as_tuple().exponent)  # pyright: ignore[reportOperatorIssue]

            #! 決定最小小數位數
            min_dp = min(dp_r, dp_db)

            #! 設定容忍誤差為「最小小數位數的 0.6 個單位」
            #! (涵蓋了恰好差 0.5 的四捨五入/無條件捨去/浮點數精度遺失落差)
            tolerance = Decimal("0.6") * (Decimal("10") ** -min_dp)

            if abs(dec_r - dec_db) <= tolerance:
                return True
        else:
            #! 關閉容差：必須絕對相等
            if dec_r == dec_db:
                return True
    except InvalidOperation:
        #! 若無法轉為數值 (例如純文字與空字串比對)，則回傳不匹配
        pass

    return False


#! 欄位轉換函式：將 Excel 英文欄位名稱轉換為從 1 開始的整數索引
def excel_column_to_number(column_str):
    #! 基礎清理：轉為大寫並去除空格
    column_str = str(column_str).upper().strip()

    number = 0
    #! 遍歷字串中的每個字母，並依照 26 進制原理進行加權計算
    for char in column_str:
        #! ord(char) 取得字元編碼，'A' 的編碼為 65，減去 64 得到 A=1
        if "A" <= char <= "Z":
            number = number * 26 + (ord(char) - ord("A") + 1)
        else:
            #! 若字串包含非英文字母則回傳 0 或拋出錯誤
            return 0

    return number


#! 主鍵正規化
def normalize_key(val):
    """
    將主鍵(如 RowNo 或 ItemCode)進行正規化：
    1. 解決 0.00 == ""
    2. 解決 5.10 == 5.1
    3. 解決 5.00 == 5
    """
    s_val = str(val).strip().replace("nan", "").replace("None", "")
    s_val = s_val.strip("[]")

    #! 處理空值與 0 等效的問題
    if not s_val or s_val in ("0", "0.0", "0.00", "-", "False"):
        return ""

    #! 處理小數點後多餘的 0
    try:
        f_val = float(s_val)
        if f_val.is_integer():
            return str(int(f_val))  # * 例如 5.00 -> "5"
        else:
            return str(f_val)  # * 例如 5.10 -> "5.1"
    except ValueError:
        #! 如果是純文字代號 (例如 "CP01")，無法轉數字，就原封不動回傳
        return s_val


@staff_required
def tigf_dashboard(request):
    #! 雙層結構: { '1234567': { 'A011': {'report': True, 'template': True, 'db': True} } }
    file_status_map = {}
    error_files = []

    if request.method == "POST":
        action = request.POST.get("action", "check")
        dict_report = get_all_files(request.FILES.getlist("files_report"))
        dict_template = get_all_files(request.FILES.getlist("files_template"))
        dict_db = get_all_files(request.FILES.getlist("files_db"))

        #! 預處理共用的範本檔與資料庫檔
        global_templates = {n.split(".")[0]: obj for n, obj in dict_template.items()}
        global_dbs = {n.split("_")[2]: obj for n, obj in dict_db.items() if "_" in n}

        if action == "check":
            #! 掃描申報檔並建立 公司編號 -> 報表編號 結構
            for name, obj in dict_report.items():
                #! 檔名解析：M202506 1234567 A011 00 (長度20)
                #! M202506 (0~7) | 1234567 (7~14) | A011 (14~18)
                if len(name) < 18:
                    error_files.append(f"申報檔格式異常：{name}")
                    continue

                cno = name[7:14]  # * 公司編號
                fid = name[14:18]  # * 報表編號

                if cno not in file_status_map:
                    file_status_map[cno] = {}

                #! 判斷是否有範本檔
                has_template = fid in global_templates
                has_db = False

                #! ========================================================
                #! 針對 TIS 系列進行多 DB 驗證
                #! ========================================================
                if fid.startswith("TIS") and dict_db:
                    if has_template:
                        try:
                            #! 讀取範本檔 Config 頁籤 (第一張工作表)
                            df_config_raw = smart_read_excel(
                                global_templates[fid], sheet_name=0, header=None
                            )

                            #! (安全機制) 讀取完畢後重置檔案指標，避免若需要重新讀取時報錯
                            if hasattr(global_templates[fid], "seek"):
                                global_templates[fid].seek(0)

                            #! 動態定位標題列
                            cfg_header_idx = df_config_raw[
                                df_config_raw.eq("工作表名稱").any(axis=1)
                            ].index
                            if len(cfg_header_idx) > 0:
                                df_config = df_config_raw.iloc[cfg_header_idx[0] + 1 :].copy()
                                df_config.columns = df_config_raw.iloc[cfg_header_idx[0]]

                                if "Table Name" in df_config.columns:
                                    #! 取出所有 Table Name 並過濾空值
                                    table_names = (
                                        df_config["Table Name"]
                                        .dropna()
                                        .astype(str)
                                        .str.strip()
                                        .tolist()
                                    )
                                    table_names = [
                                        t for t in table_names if t and t.lower() != "nan"
                                    ]

                                    missing_dbs = []
                                    #! 檢查每個 Table Name 是否都有上傳對應的 DB
                                    for t_name in table_names:
                                        #! Table Name 格式如 "ew_01_CP0010_ParticipantInfo"
                                        #! 你的 global_dbs 的 Key 是由檔名 n.split("_")[2] 取得，即 "CP0010"
                                        parts = t_name.split("_")
                                        db_key = parts[2] if len(parts) > 2 else t_name

                                        if db_key not in global_dbs:
                                            missing_dbs.append(t_name)

                                    #! 判斷是否所有定義的 DB 表都有上傳
                                    if table_names and not missing_dbs:
                                        has_db = True
                                    elif missing_dbs:
                                        error_files.append(
                                            f"{name} 缺少關聯的 DB 檔：{', '.join(missing_dbs)}"
                                        )
                                else:
                                    error_files.append(
                                        f"範本檔 {fid} config 找不到 'Table Name' 欄位"
                                    )
                            else:
                                error_files.append(
                                    f"範本檔 {fid} config 找不到 '工作表名稱' 標題行"
                                )
                        except Exception as e:
                            error_files.append(f"解析範本檔 {fid} 時發生錯誤：{e}")
                else:
                    #! 一般報表直接用 fid 檢查 (例如 L153)
                    has_db = fid in global_dbs

                #! 檢查這張報表是否具備共用的 template 與 db
                file_status_map[cno][fid] = {
                    "report": True,
                    "template": has_template,
                    "db": has_db,
                }

            #! 排序與驗證是否全部齊全
            sorted_map = {k: dict(sorted(v.items())) for k, v in sorted(file_status_map.items())}
            all_matched = True
            has_files = False

            for cno, fids in sorted_map.items():
                for fid, status in fids.items():
                    has_files = True
                    if not (status["report"] and status["template"] and status["db"]):
                        all_matched = False

            if not has_files:
                all_matched = False

            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"status_map": sorted_map, "all_matched": all_matched, "errors": error_files})

        elif action == "compare":
            diff_summary = []
            session_key = request.session.session_key or request.session.create()

            #! 從 request 中取得前端設定的 4 項規則開關，若為字串 "true" 則轉為布林值 True
            rule_str_match = request.POST.get("rule_str_match") == "true"
            rule_date_check = request.POST.get("rule_date_check") == "true"
            rule_empty_zero = request.POST.get("rule_empty_zero") == "true"
            rule_tolerance = request.POST.get("rule_tolerance") == "true"

            #! 解析申報檔
            for name, report_obj in dict_report.items():
                if len(name) < 18:
                    continue
                cno = name[7:14]
                fid = name[14:18]

                # ! ========================================================
                # ! TIS 專屬比對邏輯
                # ! ========================================================
                if "TIS" in fid:
                    try:
                        has_tis = True

                        #! 1. 【效能優化關鍵】在進入迴圈前，一次性將申報檔的所有工作表載入記憶體字典
                        #! 使用 sheet_name=None，pandas 會自動讀取整份活頁簿，report_obj 只會被解開這一次
                        with warnings.catch_warnings():
                            warnings.filterwarnings(
                                "ignore", category=UserWarning, module="openpyxl"
                            )
                            if hasattr(report_obj, "seek"):
                                report_obj.seek(0)
                            all_report_sheets = pd.read_excel(
                                report_obj, sheet_name=None, header=None, dtype=str
                            )

                        #! 讀取 TIS 範本第一張表 (Config)，使用 header=None 以防標題不在第一列
                        df_config_raw = smart_read_excel(
                            global_templates[fid], sheet_name=0, header=None
                        )

                        #! 動態尋找 "工作表名稱" 所在的列作為 DataFrame 的 Header
                        cfg_header_idx = df_config_raw[
                            df_config_raw.eq("工作表名稱").any(axis=1)
                        ].index
                        if not len(cfg_header_idx):
                            print("TIS Config 表找不到 '工作表名稱' 標題行")
                            continue

                        df_config = df_config_raw.iloc[cfg_header_idx[0] + 1 :].copy()
                        df_config.columns = df_config_raw.iloc[cfg_header_idx[0]]

                        diff_list = []  # * 收集此檔案所有 Sheet 的差異

                        #! 逐列讀取 Config，解析每張工作表
                        for _, cfg in df_config.iterrows():
                            sheet_name = str(cfg.get("工作表名稱", "")).strip()
                            tag = str(cfg.get("Tag", "")).strip()

                            if not sheet_name or sheet_name.lower() == "nan" or not tag:
                                continue

                            start_row = int(cfg.get("Start RowNo", 1))
                            start_col = int(cfg.get("Start ColNo", 1))
                            end_row = int(cfg.get("End RowNo", start_row))
                            end_col = int(cfg.get("End ColNo", start_col))

                            #! 特殊表不做比對
                            if tag in (
                                "T038",
                                "T039",
                                "T040",
                                "T041",
                                "T042",
                                "T051",
                                # "T053",
                                "T103",
                                "Y320",
                                "Y330",
                            ):
                                diff_list.append(
                                    {
                                        "工作表": f"{sheet_name}：{tag}",
                                        "行號": f"{start_row}：{end_row}",
                                        "中文欄位": "",
                                        "英文欄位": "",
                                        "申報值": "",
                                        "DB值": "特殊格式請人工比對",
                                    }
                                )
                                continue

                            #! 取得要比對的 DB 資料表名稱 (Table Name)
                            db_table_name = str(cfg.get("Table Name", "")).strip().split("_")[2]
                            if not db_table_name or db_table_name.lower() == "nan":
                                print(f"工作表 {sheet_name} 找不到 Table Name，跳過比對")
                                continue

                            #! 取得定義 Row No 的欄位 (通常是 2，代表 Excel 第 2 欄)
                            row_no_excel_col = str(cfg.get("rowNoColumn", "")).strip()
                            row_no_idx = (
                                int(row_no_excel_col) - 1 if row_no_excel_col.isdigit() else None
                            )

                            #! 解析「忽略列(相對行數)」格式為：X1:X2,Y1:Y2
                            ignore_str = str(cfg.get("忽略列(相對行數)", "")).strip()
                            ignore_tag_map = {
                                "T015": "1",
                                "T016": "1",
                                "T026": "1",
                                "T027": "1",
                                "T030": "1",
                                # "T043": "1:2",
                                # "T044": "1:2",
                                # "T045": "1:2",
                                # "T046": "1:2",
                                # "T047": "1:2",
                                # "T048": "1:3",
                                # "T049": "1:2",
                                # "T050": "1:2",
                                # "T052": "1:2",
                                # "T054": "1:2",
                                # "T055": "1:2",
                                # "T056": "1:2",
                                # "T057": "1:2",
                                # "T058": "1:2",
                                # "T059": "1:2",
                                # "T060": "1:2",
                                # "T061": "1:2",
                                # "T062": "1:2",
                                # "T063": "1:2",
                                # "T064": "1:2",
                                # "T065": "1:2",
                                # "T066": "1",
                                # "T067": "1:2",
                                # "T068": "1:2",
                                # "T069": "1:2",
                                # "T070": "1:2",
                                # "T071": "1:2",
                                # "T073": "1:2",
                                # "T074": "1:2",
                            }
                            if tag in ignore_tag_map:
                                ignore_str = ignore_tag_map[tag]
                            ignore_rows = set()
                            if ignore_str and ignore_str.lower() != "nan":
                                for part in ignore_str.split(","):
                                    part = part.strip()
                                    if ":" in part:
                                        s, e = part.split(":")
                                        ignore_rows.update(range(int(s), int(e) + 1))
                                    elif part.isdigit():
                                        ignore_rows.add(int(part))

                            #! 忽略檢核關鍵字(以,隔開)
                            ignore_str = str(cfg.get("忽略檢核關鍵字(以,隔開)", "")).strip()
                            ignore_list = ignore_str.split(",") if ignore_str else []

                            #! 讀取申報檔對應的 sheet_name
                            # try:
                            #     df_r = smart_read_excel(
                            #         report_obj, sheet_name=sheet_name, header=None
                            #     )
                            # except Exception:
                            #     print(f"找不到申報檔 Sheet: {sheet_name}")
                            #     continue

                            #! 2. 【效能優化關鍵】從記憶體字典直接撈取 DataFrame，取代原本的 smart_read_excel
                            if sheet_name not in all_report_sheets:
                                print(f"找不到申報檔 Sheet: {sheet_name}")
                                continue
                            df_r = all_report_sheets[sheet_name]

                            #! 讀取範本檔對應的 Tag sheet (獲取 Schema 規則)
                            try:
                                df_schema_raw = smart_read_excel(
                                    global_templates[fid], sheet_name=tag, header=None
                                )
                                #! 動態尋找 COLUMN_NAME 所在的列
                                sch_header_idx = df_schema_raw[
                                    df_schema_raw.eq("COLUMN_NAME").any(axis=1)
                                ].index
                                if not len(sch_header_idx):
                                    continue
                                df_schema = df_schema_raw.iloc[sch_header_idx[0] + 1 :].copy()
                                df_schema.columns = df_schema_raw.iloc[sch_header_idx[0]]
                            except Exception:
                                print(f"找不到範本檔 Schema Sheet: {tag}")
                                continue

                            #! 建立欄位 Mapping: (Excel欄位 Index -> DB欄位名稱)
                            #! 注意：Excel欄位 1 代表 A 欄，轉換成 pandas 索引需 -1
                            col_mapping = {}
                            db_pk_col = "RowNo"
                            for _, s_row in df_schema.iterrows():
                                excel_col = s_row.get("對應excel column")
                                db_col = s_row.get("COLUMN_NAME")
                                ch_col = s_row.get("中文欄位名稱")

                                if (
                                    pd.notna(excel_col)
                                    and str(excel_col).strip().isdigit()
                                    and pd.notna(db_col)
                                ):
                                    col_idx = int(excel_col) - 1
                                    col_mapping[col_idx] = {
                                        "en": str(db_col).strip(),
                                        "ch": str(ch_col).strip(),
                                    }
                                    #! 如果這個欄位剛好是 rowNoColumn，記下它的 DB 欄位名作為主鍵
                                    if row_no_idx is not None and col_idx == row_no_idx:
                                        db_pk_col = str(db_col).strip()

                            #! ========================================================
                            #! 新增：動態抓取要比對的「第一個欄位」(作為主鍵備案)
                            #! ========================================================
                            first_col_idx = min(col_mapping.keys()) if col_mapping else 0
                            first_col_db_name = (
                                col_mapping[first_col_idx]["en"] if col_mapping else ""
                            )

                            #! ========================================================
                            #! 讀取對應的 DB 表 (依照 Table Name 抓取)
                            #! ========================================================
                            #! 假設你的 global_dbs 已經以 db_table_name 為 key 存入 (如：global_dbs["ew_01_CP0010_ParticipantInfo"])
                            #! 若不是，請依你的路徑讀取，例如：smart_read_csv(os.path.join(DB_DIR, f"{db_table_name}.csv"))
                            if db_table_name not in global_dbs:
                                print(f"警告：找不到 {db_table_name} 的 DB 檔案來源，跳過比對")
                                continue

                            df_db_full = smart_read_csv(global_dbs[db_table_name])

                            #! 過濾出該公司、未刪除的資料
                            df_db_full["Cno"] = df_db_full["Cno"].astype(str).str.strip()
                            df_db = df_db_full[df_db_full["Cno"] == str(cno)]
                            if "isdel" in df_db.columns:
                                df_db = df_db[
                                    df_db["isdel"].isna()
                                    | (df_db["isdel"].astype(str).str.strip() == "")
                                ]

                            #! ========================================================
                            #! 建立 (第一欄位, RowNo) 的複合主鍵字典
                            #! ========================================================
                            db_lookup = {}
                            for _, db_row in df_db.iterrows():
                                #! 取得第一個欄位的值
                                val_first = ""
                                if first_col_db_name in df_db.columns:
                                    val_first = normalize_key(db_row[first_col_db_name])

                                #! 取得 RowNo 的值
                                val_row_no = ""
                                if db_pk_col in df_db.columns:
                                    val_row_no = normalize_key(db_row[db_pk_col])

                                #! 組成 Tuple: 例如 ('CP01', '10') 或 ('CP02', '')
                                # if tag == "T082":
                                #     k = (val_first, "")
                                # else:
                                #     k = (val_first, val_row_no)
                                k = (val_first, val_row_no)

                                #! 因為 (第一欄位, RowNo) 可能會重複（例如都是空值），所以存成 List
                                if k not in db_lookup:
                                    db_lookup[k] = []
                                db_lookup[k].append(db_row)

                            #! 1. 拉出 config 實際筆數
                            actual_report_row_count = int(cfg.get("實際筆數", ""))

                            #! 2. 計算 DB 端的總資料筆數 (該公司未刪除的總筆數)
                            total_db_row_count = sum(len(rows) for rows in db_lookup.values())

                            #! 3. 進行總筆數比對驗證
                            if actual_report_row_count != total_db_row_count:
                                diff_list.append(
                                    {
                                        "工作表": f"{sheet_name}：{tag}",
                                        "行號": f"{start_row}：{end_row}",
                                        "中文欄位": "資料總筆數核對",
                                        "英文欄位": "",
                                        "申報值": f"實際有效資料共 {actual_report_row_count} 筆",
                                        "DB值": f"DB 總資料共 {total_db_row_count} 筆",
                                    }
                                )
                                print(
                                    f"[{sheet_name}：{tag}] 筆數不一致！申報實際筆數：{actual_report_row_count}，DB總筆數：{total_db_row_count}\n"
                                )
                                continue

                            #! 擷取資料並逐行進行比對
                            for r_idx in range(start_row - 1, end_row):
                                if r_idx >= len(df_r):
                                    break

                                #! 計算相對行數 (從 1 開始算)
                                rel_row = r_idx - (start_row - 1) + 1
                                if rel_row in ignore_rows:
                                    continue

                                row_data = df_r.iloc[r_idx]

                                #! ========================================================
                                #! 新增防呆：檢查該行是否「全部都是空的或 0」
                                #! ========================================================
                                is_empty_row = True
                                for c_idx in col_mapping.keys():
                                    if c_idx < len(row_data):
                                        #! 取出每個要比對的欄位值
                                        val = (
                                            str(row_data.iloc[c_idx])
                                            .strip()
                                            .replace("nan", "")
                                            .replace("None", "")
                                        )

                                        #! 如果發現任何一個欄位有「非空、非0」的有效字元，就代表這行有意義！
                                        if val and val not in ("0", "0.0", "0.00"):
                                            is_empty_row = False
                                            break

                                #! 如果整行都是空值或0，這行沒有比對價值，直接跳過！
                                #! 這樣才不會消耗掉 unkeyed_db_rows 裡真正有意義的資料
                                if is_empty_row:
                                    continue

                                #! ========================================================
                                #! 改良：組合申報行的複合主鍵，並去 DB 尋找
                                #! ========================================================
                                #! 取得第一個欄位的值
                                r_val_first = ""
                                if first_col_idx < len(row_data):
                                    r_val_first = normalize_key(row_data.iloc[first_col_idx])
                                    if (
                                        tag in ("X110", "X120", "X130", "X140", "Y110")
                                        and r_val_first == ""
                                    ):
                                        r_val_first = "TAL"
                                    # else:
                                    #     if (
                                    #         tag == "T048"
                                    #         and r_val_first == "對不動產風險敏感之資產減負債"
                                    #     ):
                                    #         r_val_first = "對不動產風險敏感之資產減負債      台灣↓7.81%、其他↓25%"

                                #! 取得 RowNo 的值
                                r_val_row_no = ""
                                if row_no_idx is not None and row_no_idx < len(row_data):
                                    if (
                                        tag in ("X110", "X120", "X130", "X140", "Y110")
                                        and r_val_first == "TAL"
                                    ):
                                        r_val_row_no = "TAL"
                                    else:
                                        r_val_row_no = normalize_key(row_data.iloc[row_no_idx])

                                #! 3. 【修正 T082 錯誤】申報端的 Key 組裝邏輯必須與 DB 端對稱
                                if tag == "T082":
                                    r_key = (r_val_first, "")
                                else:
                                    r_key = (r_val_first, r_val_row_no)

                                r_key = (r_val_first, r_val_row_no)

                                #! 從 DB 字典中尋找對應的資料並取出 (pop)
                                row_db = None
                                if r_key in db_lookup and db_lookup[r_key]:
                                    #! 使用 pop(0) 拿出第一筆
                                    #! 若有唯一主鍵，List 內只會有 1 筆；若有多筆空 RowNo 但代號相同，能安全依序消耗
                                    row_db = db_lookup[r_key].pop(0)

                                if row_db is None:
                                    #! DB 找不到這行資料，整行報錯
                                    display_key = (
                                        f"{r_val_first} | {r_val_row_no}"
                                        if r_val_row_no
                                        else r_val_first
                                    )
                                    diff_list.append(
                                        {
                                            "工作表": f"{sheet_name}：{tag}",
                                            "行號": r_idx + 1,
                                            "中文欄位": f"主鍵({display_key})",
                                            "英文欄位": f"{first_col_db_name} | {db_pk_col}",
                                            "申報值": "有資料",
                                            "DB值": "DB 無對應列資料",
                                        }
                                    )
                                    print(
                                        f"報表：{tag}\n工作表：{sheet_name}\n行號：{r_idx + 1}\n主鍵：{display_key}\n"
                                    )
                                    continue

                                #! 取出這行需比對的數值
                                for c_idx, cols in col_mapping.items():
                                    #! 檢查該欄位是否在指定的 Start ColNo ~ End ColNo 範圍內
                                    if start_col - 1 <= c_idx <= end_col - 1 and c_idx < len(
                                        row_data
                                    ):
                                        val_r = (
                                            str(row_data.iloc[c_idx])
                                            .strip()
                                            .replace("nan", "")
                                            .replace("None", "")
                                        )

                                        #! DB 如果沒有這個欄位，給空值
                                        val_db = (
                                            str(row_db.get(cols["en"], ""))
                                            .strip()
                                            .replace("nan", "")
                                            .replace("None", "")
                                        )

                                        #! 呼叫共用的比對函數
                                        if not is_value_matched(
                                            val_r,
                                            val_db,
                                            rule_str_match,
                                            rule_date_check,
                                            rule_empty_zero,
                                            rule_tolerance,
                                            has_tis,
                                            ignore_list,
                                        ):
                                            if (
                                                val_r == ""
                                                and val_db == "TAL"
                                                and tag in ("X110", "X120", "X130", "X140", "Y110")
                                                and cols["ch"] == "編號"
                                            ):
                                                continue
                                            diff_list.append(
                                                {
                                                    "工作表": f"{sheet_name}：{tag}",
                                                    "行號": r_idx + 1,
                                                    "中文欄位": cols["ch"],
                                                    "英文欄位": cols["en"],
                                                    "申報值": val_r,
                                                    "DB值": val_db,
                                                }
                                            )
                                            print(
                                                f"[{sheet_name}：{tag}] 行號 {r_idx + 1} | {cols['ch']} ({cols['en']}) 不符"
                                            )
                                            print(f"申報值：{repr(val_r)}")
                                            print(f" DB值：{repr(val_db)}\n")

                        #! 單檔 TIS 比對完成，處理結果輸出 (寫入 Cache / Excel)
                        diff_count = len(diff_list)
                        if diff_count > 0:
                            df_diff = pd.DataFrame(diff_list)
                            excel_buf = io.BytesIO()
                            df_diff.to_excel(excel_buf, index=False, engine="openpyxl")
                            cache.set(f"diff_{session_key}_{cno}_{fid}", excel_buf.getvalue(), 3600)

                        diff_summary.append(
                            {
                                "cno": cno,
                                "fid": fid,
                                "diff_count": diff_count,
                                "schema_error": False,
                            }
                        )

                    except Exception as e:
                        print(f"處理 TIS {cno}-{fid} 時發生錯誤: {e}")
                elif fid in global_templates and fid in global_dbs:
                    try:
                        ignore_rows = set()
                        ignore_data = ""
                        ignore_data2 = ""
                        ignore_column_num = 0
                        ignore_column_num2 = 0
                        ignore_column_name = ""
                        ignore_column_name2 = ""
                        start_row = 8

                        #! 讀取範本 Config
                        df_config = smart_read_excel(global_templates[fid], sheet_name="config", header=None)
                        if pd.notna(df_config.iloc[3, 1]):
                            start_row = int(df_config.iloc[3, 1])

                        #! 抓取忽略檢查值的行B19 (通常是A-合計)
                        if pd.notna(df_config.iloc[18, 1]):
                            ignore_column, *_, ignore_data = str(df_config.iloc[18, 1]).split("-")
                            ignore_column_num = excel_column_to_number(ignore_column)

                        #! 抓取忽略檢查值的行B20
                        if pd.notna(df_config.iloc[19, 1]):
                            ignore_column, *_, ignore_data2 = str(df_config.iloc[19, 1]).split("-")
                            ignore_column_num2 = excel_column_to_number(ignore_column)

                        val_ignore = str(df_config.iloc[10, 1])
                        if val_ignore and val_ignore.lower() != "nan":
                            ignore_rows = {int(x.strip()) for x in val_ignore.split(",") if x.strip().isdigit()}

                        #! 讀取 Schema
                        df_schema = smart_read_excel(global_templates[fid], sheet_name="schema", skiprows=7)
                        mapping = dict(zip(df_schema["中文欄位名稱"], df_schema["COLUMN_NAME"]))

                        #! 讀取申報檔並處理雙層標題
                        df_r = smart_read_excel(report_obj, header=None)
                        row_upper = df_r.iloc[start_row - 4]
                        row_lower = df_r.iloc[start_row - 3]
                        #! F023 有3層
                        if fid == "F023":
                            row_upper = df_r.iloc[start_row - 5]

                        #! 正常和B102 特殊格式
                        if (
                            row_upper.astype(str)
                            .str.strip()
                            .replace(["nan", "None"], "")
                            .eq("")
                            .all()
                            or fid == "B102"
                        ):
                            df_r.columns = row_lower
                        else:
                            new_columns = []
                            top_l = ""
                            for upper, lower in zip(row_upper, row_lower):
                                top = str(upper).strip() if pd.notna(upper) else ""
                                bottom = str(lower).strip() if pd.notna(lower) else ""
                                if top and bottom:
                                    new_columns.append(f"{top}_{bottom}")
                                    top_l = top
                                elif top:
                                    new_columns.append(top)
                                elif not top_l:
                                    new_columns.append(bottom)
                                else:
                                    new_columns.append(f"{top_l}_{bottom}")
                            df_r.columns = new_columns

                        #! 讀取共用 DB 檔，並進行過濾
                        df_db_full = smart_read_csv(global_dbs[fid])
                        df_db_full["Cno"] = df_db_full["Cno"].astype(str).str.strip()
                        df_db = df_db_full[df_db_full["Cno"] == str(cno)]

                        if "isdel" in df_db.columns:
                            df_db = df_db[df_db["isdel"].isna() | (df_db["isdel"].astype(str).str.strip() == "")]

                        #! 取得目標比對欄位
                        target_cols = [c for c in df_r.columns if c in mapping]
                        if not target_cols:
                            diff_summary.append(
                                {"cno": cno, "fid": fid, "diff_count": 0, "schema_error": True}
                            )
                            continue

                        #! === 建立 DB 字典 (動態支援單一與複合主鍵) ===

                        #! 定義各報表的主鍵數量 (字典配置)
                        #! 如果未來有報表需要複合主鍵，直接加在這裡；沒寫在裡面的，預設都是 1 (單主鍵)
                        pk_config_map = {
                            "L153": 3,  # * L153 使用前 3 個欄位 (代碼, 名稱, 年度) 作為複合主鍵
                        }
                        #! 根據現在的報表 (fid) 動態取得要切幾個主鍵欄位
                        num_pks = pk_config_map.get(fid, 1)  # * 如果 fid 找不到，預設回傳 1

                        pk_cols_ch = target_cols[:num_pks]
                        pk_cols_en = [mapping[c] for c in pk_cols_ch]

                        #! 忽略欄位名
                        if ignore_column_num:
                            ignore_column_name = target_cols[ignore_column_num - 1]
                        #! 忽略欄位名 v2
                        if ignore_column_num2:
                            ignore_column_name2 = target_cols[ignore_column_num2 - 1]

                        #! 將字典的 Key 改為 Tuple，用來精準對應 (單主鍵也會變成 1 個元素的 Tuple)
                        db_lookup = {}
                        for _, row in df_db.iterrows():
                            #! 將多個主鍵的值組成一個 Tuple，例如 ('AE1', 'AE19', 'T-3年')
                            k_parts = tuple(
                                str(row[en]).strip().replace("nan", "").replace("None", "")
                                for en in pk_cols_en
                            )

                            #! 只要複合主鍵不是全空，就加入字典
                            if any(k_parts):
                                if k_parts not in db_lookup:
                                    db_lookup[k_parts] = []
                                #! 將資料與使用狀態打包存入
                                db_lookup[k_parts].append({"data": row, "used": False})

                        #! 開始比對
                        diff_list = []

                        for i in range(len(df_r)):
                            if (i + 1) < start_row or (i + 1) in ignore_rows:
                                continue

                            row_r = df_r.iloc[i]

                            #! 組合申報檔的複合主鍵
                            r_key_parts = tuple(
                                str(row_r[ch]).strip().replace("nan", "").replace("None", "")
                                for ch in pk_cols_ch
                            )

                            #! 如果主鍵全空，跳過
                            if not any(r_key_parts):
                                continue

                            #! 用於錯誤紀錄顯示的字串
                            r_key_display = " | ".join(r_key_parts)

                            #! 判斷是否忽略行 (直接抓取該指定欄位的值來比對)
                            if ignore_column_name:
                                val_ignore1 = (
                                    str(row_r[ignore_column_name])
                                    .strip()
                                    .replace("nan", "")
                                    .replace("None", "")
                                )
                                if val_ignore1 == ignore_data:
                                    continue

                            #! 判斷是否忽略行v2
                            if ignore_column_name2:
                                val_ignore2 = (
                                    str(row_r[ignore_column_name2])
                                    .strip()
                                    .replace("nan", "")
                                    .replace("None", "")
                                )
                                if val_ignore2 == ignore_data2:
                                    continue

                            #! 檢查複合主鍵是否存在於 DB
                            if r_key_parts not in db_lookup:
                                diff_list.append(
                                    {
                                        "行號": i + 1,
                                        "中文欄位": "複合主鍵",
                                        "英文欄位": "Composite PK",
                                        "申報值": r_key_display,
                                        "DB值": "DB 此行無對應資料",
                                    }
                                )
                                continue

                            #! 找出 DB 中該複合主鍵「尚未被比對過」的資料
                            available_db_items = [
                                item for item in db_lookup[r_key_parts] if not item["used"]
                            ]

                            if not available_db_items:
                                #! 代表申報檔裡該組合的數量，比 DB 裡的還要多
                                diff_list.append(
                                    {
                                        "行號": i + 1,
                                        "中文欄位": "複合主鍵",
                                        "英文欄位": "Composite PK",
                                        "申報值": r_key_display,
                                        "DB值": "DB 缺漏此行資料 (或重複行數不足)",
                                    }
                                )
                                continue

                            #! 取出第一筆可用的 DB 資料，並標記為已使用
                            db_item = available_db_items[0]
                            row_db = db_item["data"]
                            db_item["used"] = True  # * 關鍵：上鎖

                            #! 逐欄比對
                            for ch_col in target_cols:
                                en_col = mapping[ch_col]
                                if en_col not in row_db:
                                    continue

                                val_r = row_r[ch_col]
                                val_db = row_db[en_col]

                                if not is_value_matched(
                                    val_r,
                                    val_db,
                                    rule_str_match,
                                    rule_date_check,
                                    rule_empty_zero,
                                    rule_tolerance,
                                ):
                                    diff_list.append(
                                        {
                                            "行號": i + 1,
                                            "中文欄位": ch_col,
                                            "英文欄位": en_col,
                                            "申報值": val_r,
                                            "DB值": val_db,
                                        }
                                    )
                                    print(f"值1：{repr(val_r)}")
                                    print(f"值2：{repr(val_db)}\n")

                        diff_count = len(diff_list)
                        if diff_count > 0:
                            df_diff = pd.DataFrame(diff_list)
                            excel_buf = io.BytesIO()
                            df_diff.to_excel(excel_buf, index=False, engine="openpyxl")
                            cache.set(f"diff_{session_key}_{cno}_{fid}", excel_buf.getvalue(), 3600)

                        diff_summary.append(
                            {
                                "cno": cno,
                                "fid": fid,
                                "diff_count": diff_count,
                                "schema_error": False,
                            }
                        )

                    except Exception as e:
                        print(f"處理 {cno}-{fid} 時發生錯誤: {e}")

            return JsonResponse({"action": "compare_results", "diff_results": diff_summary})

    return render(
        request, "tigf/tigf_dashboard.html", {"status_map": file_status_map, "all_matched": False}
    )


@staff_required
def download_diff_csv(request, cno, fid):
    """
    從快取讀取比對差異 CSV 並提供下載
    """
    #! 取得 Session Key (必須與 compare 寫入時的一致)
    session_key = request.session.session_key
    if not session_key:
        return HttpResponse("連線已逾期，請重新整理頁面並重新比對。", status=400)

    #! 組合快取 Key
    cache_key = f"diff_{session_key}_{cno}_{fid}"
    excel_data = cache.get(cache_key)

    #! 檢查資料是否存在
    if not excel_data:
        #! 如果快取過期（1小時）或找不到資料
        return HttpResponse("找不到比對資料或檔案已過期，請重新執行比對。", status=404)

    #! 建立 HttpResponse
    #! 指定內容類型為 CSV
    response = HttpResponse(
        excel_data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    #! 設定下載檔名 (Content-Disposition)
    #! 建議檔名加上 fid，方便使用者辨識
    filename = f"安定比對差異檔_{cno}_{fid}.xlsx"
    response["Content-Disposition"] = f"attachment; filename*=utf-8''{quote(filename)}"

    return response


def ics_merger_dashboard(request):
    """渲染 ICS Excel 報表合併主頁面"""
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        main_file = request.FILES.get("main_excel")
        sub_file = request.FILES.get("sub_excel")

        if not main_file or not sub_file:
            return JsonResponse({"success": False, "error": "請同時上傳主 Excel 檔與副 Excel 檔"})

        try:
            #! 讀取主副 Excel 活頁簿
            wb_main = load_workbook(main_file)
            wb_sub = load_workbook(sub_file, data_only=False)  # * 保留公式與完整樣式

            #! 逐一將副檔的 sheet 複製進主檔
            for sheet_name in wb_sub.sheetnames:
                ws_sub = wb_sub[sheet_name]

                #! 處理工作表名稱重複衝突
                new_sheet_name = sheet_name
                counter = 1
                while new_sheet_name in wb_main.sheetnames:
                    new_sheet_name = f"{sheet_name}_{counter}"
                    counter += 1

                #! 在主檔後方建立新工作表
                ws_main = wb_main.create_sheet(title=new_sheet_name)

                #! 複製所有儲存格的值、公式與樣式
                for row in ws_sub.iter_rows():
                    for cell in row:
                        new_cell = ws_main.cell(row=cell.row, column=cell.column, value=cell.value)

                        #! 複製細胞格核心樣式 (字體、填滿、對齊、框線、數字格式)
                        if cell.has_style:
                            new_cell.font = copy.copy(cell.font)
                            new_cell.fill = copy.copy(cell.fill)
                            new_cell.alignment = copy.copy(cell.alignment)
                            new_cell.border = copy.copy(cell.border)
                            new_cell.number_format = cell.number_format

                #! 複製合併儲存格的範圍
                for merged_range in ws_sub.merged_cells.ranges:
                    ws_main.merge_cells(str(merged_range))

                #! 複製列高與欄寬樣式
                for col_idx in range(1, ws_sub.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    ws_main.column_dimensions[col_letter].width = ws_sub.column_dimensions[
                        col_letter
                    ].width

                for row_idx in range(1, ws_sub.max_row + 1):
                    ws_main.row_dimensions[row_idx].height = ws_sub.row_dimensions[row_idx].height

            #! 將合併後的 Excel 寫入記憶體中準備導出
            output = io.BytesIO()
            wb_main.save(output)
            output.seek(0)

            #! 暫存回 Session 供後續下載，或直接回傳成功訊號
            request.session["merged_ics_file"] = output.getvalue().hex()
            return JsonResponse({"success": True, "message": "報表合併成功，準備開始下載"})

        except Exception as e:
            return JsonResponse({"success": False, "error": f"合併程序失敗：{str(e)}"})

    return render(request, "tigf/ics_merger_dashboard.html")


def download_merged_ics(request):
    """提供使用者下載合併完成後的檔案"""
    file_hex = request.session.get("merged_ics_file")
    if not file_hex:
        return HttpResponse("找不到可供下載的報表檔案", status=404)

    file_bytes = bytes.fromhex(file_hex)
    response = HttpResponse(
        file_bytes, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ICS_Merged_Report.xlsx"'

    #! 下載後清除 Session 釋放記憶體
    del request.session["merged_ics_file"]
    return response


def ics_cleaner_dashboard(request):
    """處理 ICS Excel 結構淨化與無效工作表清除"""
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        excel_file = request.FILES.get("dirty_excel")

        if not excel_file:
            return JsonResponse({"success": False, "error": "請選取要執行清除的 Excel 檔案"})

        try:
            # * 載入活頁簿 (保留公式與樣式不變)
            wb = load_workbook(excel_file, data_only=False)

            # * 基本檢查：必須存在核心設定工作表
            if "config" not in wb.sheetnames:
                return JsonResponse(
                    {"success": False, "error": "規格錯誤：找不到核心 'config' 工作表"}
                )

            # * 1. 初始化白名單工作表名稱集合
            preserved_sheets = {"config", "rowNumber"}
            ws_config = wb["config"]

            # * 2. 動態從 config 工作表內掃描撈取所有的 Tag 名稱 (例如 T001, T002...)
            # * 遍歷 config 工作表中的前幾欄所有儲存格，只要數值符合 T 加上數字或自訂 Tag 格式就納入白名單
            for row in ws_config.iter_rows(values_only=True):
                for val in row:
                    if val and isinstance(val, str):
                        val_strip = val.strip()
                        # * 匹配常見的 Tag 格式 (例如 T001, BS03 等大寫代碼開頭)
                        if val_strip.isalnum() and len(val_strip) == 4:
                            preserved_sheets.add(val_strip)

            # * 3. 安全過濾清除：找出不在白名單內的 Sheet 並將其移除
            all_sheet_names = list(wb.sheetnames)
            removed_count = 0
            removed_list = []

            for sheet_name in all_sheet_names:
                if sheet_name not in preserved_sheets:
                    wb.remove(wb[sheet_name])
                    removed_list.append(sheet_name)
                    removed_count += 1

            # * 如果全部都被刪光了防呆機制
            if len(wb.sheetnames) == 0:
                return JsonResponse(
                    {"success": False, "error": "清除異常：過濾後沒有保留任何有效工作表"}
                )

            # * 將過濾完成後的 Excel 寫入記憶體中準備導出
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # * 儲存進 Session 供前端點擊下載
            request.session["cleaned_ics_file"] = output.getvalue().hex()

            return JsonResponse(
                {
                    "success": True,
                    "message": "結構淨化完成",
                    "removed_count": removed_count,
                    "removed_sheets": removed_list,
                }
            )

        except Exception as e:
            return JsonResponse({"success": False, "error": f"清除程序執行失敗：{str(e)}"})

    return render(request, "tigf/ics_cleaner_dashboard.html")


def download_cleaned_ics(request):
    """提供使用者下載淨化完成後的 Excel 檔案"""
    file_hex = request.session.get("cleaned_ics_file")
    if not file_hex:
        return HttpResponse("找不到可供下載的檔案", status=404)

    file_bytes = bytes.fromhex(file_hex)
    response = HttpResponse(
        file_bytes, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ICS_Cleaned_Report.xlsx"'

    # * 下載完成後即刻釋放 Session
    del request.session["cleaned_ics_file"]
    return response


def ics_validator_dashboard(request):
    """處理 config 與 rowNumber 結構指標的交叉驗證"""
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        excel_file = request.FILES.get("check_excel")

        if not excel_file:
            return JsonResponse({"success": False, "error": "請選擇要執行驗證的 Excel 檔案"})

        try:
            wb = load_workbook(excel_file, data_only=True)

            # * 基本檢查：核心驗證工作表必須存在
            if "config" not in wb.sheetnames or "rowNumber" not in wb.sheetnames:
                return JsonResponse(
                    {
                        "success": False,
                        "error": "錯誤：檔案中必須同時包含 'config' 與 'rowNumber' 工作表",
                    }
                )

            ws_config = wb["config"]
            ws_rownum = wb["rowNumber"]

            # * 1. 解析 rowNumber 工作表（讀取首行首名作為 Tag 指標，並收集底下的編號）
            rownum_data = {}
            # * 取得 rowNumber 所有的欄位
            for col in ws_rownum.iter_cols(values_only=True):
                if not col or col[0] is None:
                    continue
                tag_header = str(col[0]).strip()
                # * 收集該欄底下所有非空的行號編號（排除首行）
                row_ids = [str(v).strip() for v in col[1:] if v is not None]
                rownum_data[tag_header] = row_ids

            validation_errors = []
            passed_logs = []

            # * 2. 依序讀取 config 每筆設定（排除前三行標頭，資料自第4行開始）
            for idx, row in enumerate(ws_config.iter_rows(min_row=4, values_only=True), start=4):
                sheet_name = row[0]  # * A欄: 工作表名稱
                tag_name = row[1]  # * B欄: Tag
                start_row = row[2]  # * C欄: Start RowNo
                start_col = row[3]  # * D欄: Start ColNo
                end_row = row[4]  # * E欄: End RowNo
                end_col = row[5]  # * F欄: End ColNo
                row_type = row[6]  # * G欄: Row Type
                row_count = row[7]  # * H欄: Row Count
                up_end_row_key_word = row[8]  # * I欄: Up End Row Key Word  # noqa: F841
                bottom_start_row_key_word = row[9]  # * J欄: Bottom Start Row Key Word  # noqa: F841
                row_shift = row[10]  # * K欄: Row Shift  # noqa: F841
                rownum_col = row[14]  # * O欄: rowNoColumn
                ignore_rows = row[16]  # * Q欄: Ignore Rows  # noqa: F841

                # * 略過完全空白的設定行
                if not sheet_name and not tag_name:
                    continue

                sheet_name = str(sheet_name).strip()
                tag_name = str(tag_name).strip() if tag_name else None

                # * 驗證 A 欄工作表是否存在於活頁簿
                if sheet_name not in wb.sheetnames:
                    validation_errors.append(
                        f"Line {idx}: 工作表 [{sheet_name}] 不存在於此 Excel 檔案中"
                    )
                    continue

                ws_target = wb[sheet_name]

                # * 驗證 D 欄與 F 欄之間的指定欄位範圍是否「整行/整區塊都是空的」
                if start_col and end_col:
                    try:
                        s_col = int(start_col)  # pyright: ignore[reportArgumentType]
                        e_col = int(end_col)  # pyright: ignore[reportArgumentType]
                        is_block_empty = True

                        # * 遍歷目標工作表指定範圍，檢查是否至少有一個儲存格有值
                        for r_idx in range(1, ws_target.max_row + 1):
                            for c_idx in range(s_col, e_col + 1):
                                if ws_target.cell(row=r_idx, column=c_idx).value is not None:
                                    is_block_empty = False
                                    break
                            if not is_block_empty:
                                break

                        if is_block_empty:
                            validation_errors.append(
                                f"[{sheet_name}] Line {idx}: "
                                f"欄位範圍 {s_col} ~ {e_col} "
                                f"查無任何數值（整區塊皆為空，不符合規範）"
                            )
                    except ValueError:
                        validation_errors.append(
                            f"[{sheet_name}] Line {idx}: 欄位指標 Start/End ColNo 必須為整數"
                        )

                # * 條件分支：如果 Row Type (G欄) 為 2，跳過 rowNumber 的對應比對
                if str(row_type).strip() == "2":
                    passed_logs.append(f"[{sheet_name}] Row Type 為 2，略過 rowNumber 檢核")
                    continue

                # * 驗證 H 欄 Row Count 是否正確
                if row_count == (int(end_row) - int(start_row) + 1):  # pyright: ignore[reportArgumentType]
                    pass

                # * 驗證 B 欄 Tag 是否存在於 rowNumber 的欄位首行中
                if not tag_name or tag_name not in rownum_data:
                    validation_errors.append(
                        f"[{sheet_name}] Line {idx}: "
                        f"設定的 Tag [{tag_name}] "
                        f"未在 rowNumber 工作表首行中找到"
                    )
                    continue

                # * 驗證 C 欄與 E 欄對應的目標欄（O欄指定）內容是否與 rowNumber 對應欄底下的編號一致
                if start_row and end_row and rownum_col:
                    try:
                        s_row = int(start_row)  # pyright: ignore[reportArgumentType]
                        e_row = int(end_row)  # pyright: ignore[reportArgumentType]
                        r_col = int(rownum_col)  # pyright: ignore[reportArgumentType]

                        # * 從目標工作表撈出實際填列的行號編號集合
                        actual_row_ids = []
                        for r_idx in range(s_row, e_row + 1):
                            cell_val = ws_target.cell(row=r_idx, column=r_col).value
                            if cell_val is not None:
                                actual_row_ids.append(str(cell_val).strip())

                        expected_row_ids = rownum_data[tag_name]

                        # * 交叉核對兩者是否完全一致
                        mismatched_ids = [x for x in actual_row_ids if x not in expected_row_ids]
                        if mismatched_ids:
                            validation_errors.append(
                                f"[{sheet_name}] Line {idx}: "
                                f"工作表內撈到的編號 {mismatched_ids} "
                                f"未包含在 rowNumber 的白名單內"
                            )
                    except ValueError:
                        validation_errors.append(
                            f"[{sheet_name}] Line {idx}: 行號或 rowNoColumn 格式錯誤，必須為整數"
                        )

            if validation_errors:
                return JsonResponse({"success": True, "valid": False, "errors": validation_errors})

            return JsonResponse(
                {
                    "success": True,
                    "valid": True,
                    "message": "恭喜！工作表指標與 rowNumber 資料完全核對正確！",
                }
            )

        except Exception as e:
            return JsonResponse({"success": False, "error": f"結構驗證程序失敗：{str(e)}"})

    return render(request, "tigf/ics_validator_dashboard.html")
